from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

import csv

def create_csv(out_file_path, days_num, headers):
    days = [day for day in range(0,days_num)]
    headers = headers

    for day in days:
        headers.append(day)

    with open(out_file_path, "w", encoding='utf-16', newline='') as file:
        writer = csv.writer(file)
        writer.writerow(headers)

    return days

def save_csv(out_file_path, headers, values):
    with open(out_file_path, 'a', newline='', encoding='utf-16') as file:
        writer = csv.DictWriter(file, fieldnames=headers)
        writer.writerow(values)

def get_activites_data(sheet,days_num):
    output_dir = 'output/act_file_5.csv'
    headers = ['Наименование работ', 'план/факт']

    days = create_csv(output_dir, days_num, headers)

    for cell in sheet['Z']:
        filter_cell = sheet.cell(row=cell.row, column=column_index_from_string('BY')).value
        work_name = sheet.cell(row=cell.row, column=column_index_from_string('L')).value
        start_column_id = column_index_from_string('AS')

        if cell.value == 'план' and type(filter_cell) == int:
            plan_list = [c.value for c in sheet[cell.row][start_column_id - 1: start_column_id + 30]]

            values_dict = {
                'Наименование работ': work_name + '_act',
                'план/факт': 'план',
            }

            plan_dict = dict(zip(days, plan_list))
            values_dict.update(plan_dict)

            save_csv(output_dir, headers, values_dict)

        if cell.value == 'факт' and type(filter_cell) == int:
            fact_list = [c.value for c in sheet[cell.row][start_column_id - 1: start_column_id + 30]]

            values_dict = {
                'Наименование работ': work_name,
                'план/факт': 'факт',
            }

            fact_dict = dict(zip(days, fact_list))
            values_dict.update(fact_dict)

            save_csv(output_dir, headers, values_dict)

def get_resources_data(sheet, days_num):
    output_dir = 'output/res_file_5.csv'
    headers = ['Ресурсы', 'Субподрядчик', 'план/факт']

    days = create_csv(output_dir,days_num, headers)

    for cell in sheet['B']:
        resource_name = sheet.cell(row=cell.row, column=column_index_from_string('A')).value
        start_column_id = column_index_from_string('G')

        if cell.row >= 4 and type(cell.value) == str:
            plan_list = [c.value for c in sheet[cell.row][start_column_id - 1: start_column_id + days_num]]
            fact_list = [c.value for c in sheet[cell.row + 1][start_column_id - 1: start_column_id + days_num]]
            values_dict = {
                'Ресурсы': resource_name + '_res',
                'Субподрядчик': cell.value,
                'план/факт': 'план',
            }

            plan_dict = dict(zip(days, plan_list))
            values_dict.update(plan_dict)

            save_csv(output_dir, headers, values_dict)

            values_dict = {
                'Ресурсы': resource_name + '_res',
                'Субподрядчик': cell.value,
                'план/факт': 'факт',
            }

            fact_dict = dict(zip(days, fact_list))
            values_dict.update(fact_dict)

            save_csv(output_dir, headers, values_dict)

def main():
    inp_file_path = 'data/5.xlsm'
    wb = load_workbook(inp_file_path, data_only=True)

    activities_sheet = wb['МСГ']
    resources_sheet = wb['Ресурсы']

    get_activites_data(activities_sheet, 30)
    get_resources_data(resources_sheet, 30)

if __name__ == "__main__":
    main()