from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

import csv

def create_csv(out_file_path, days_num, headers):
    """creates a csv file with the passed headers"""
    days = [day for day in range(1,days_num+1)]
    headers = headers

    for day in days:
        headers.append(day)

    with open(out_file_path, "w", encoding='utf-16', newline='') as file:
        writer = csv.writer(file)
        writer.writerow(headers)

    return days

def update_csv(out_file_path, headers, values):
    """updates the csv file line by line"""
    with open(out_file_path, 'a', encoding='utf-16', newline='') as file:
        writer = csv.DictWriter(file, fieldnames=headers)
        writer.writerow(values)

def get_activites_data(sheet,days_num):
    """collects data about activities"""
    output_dir = 'output/act_file_5.csv'
    headers = ['Наименование работ', 'план/факт']

    days = create_csv(output_dir, days_num, headers)

    for cell in sheet['Z']:
        filter_cell = sheet.cell(row=cell.row, column=column_index_from_string('BY')).value
        start_column_id = column_index_from_string('AS')

        if cell.value == 'план' and type(filter_cell) == int:
            work_name = sheet.cell(row=cell.row, column=column_index_from_string('L')).value
            plan_list = [c.value for c in sheet[cell.row][start_column_id - 1: start_column_id + 30]]

            plan_values_dict = {
                'Наименование работ': work_name + '_act',
                'план/факт': 'план',
            }

            plan_dict = dict(zip(days, plan_list))
            plan_values_dict.update(plan_dict)

            update_csv(output_dir, headers, plan_values_dict)

        if cell.value == 'факт' and type(filter_cell) == int:
            work_name = sheet.cell(row=cell.row - 1, column=column_index_from_string('L')).value
            fact_list = [c.value for c in sheet[cell.row][start_column_id - 1: start_column_id + 30]]

            fact_values_dict = {
                'Наименование работ': work_name + '_act',
                'план/факт': 'факт',
            }

            fact_dict = dict(zip(days, fact_list))
            fact_values_dict.update(fact_dict)

            update_csv(output_dir, headers, fact_values_dict)

def get_resources_data(sheet, days_num):
    """collects data about resources"""
    output_dir = 'output/res_file_5.csv'
    headers = ['Ресурсы', 'Субподрядчик', 'план/факт']

    days = create_csv(output_dir,days_num, headers)

    for cell in sheet['B']:
        resource_name = sheet.cell(row=cell.row, column=column_index_from_string('A')).value
        start_column_id = column_index_from_string('D')

        if cell.row >= 4 and type(cell.value) == str:
            plan_list = [c.value for c in sheet[cell.row][start_column_id - 1: start_column_id + days_num]]
            fact_list = [c.value for c in sheet[cell.row + 1][start_column_id - 1: start_column_id + days_num]]
            plan_values_dict = {
                'Ресурсы': resource_name + '_res',
                'Субподрядчик': cell.value,
                'план/факт': 'план',
            }

            plan_dict = dict(zip(days, plan_list))
            plan_values_dict.update(plan_dict)

            # только пустые значения в строке не добавляются в csv
            if set(plan_dict.values()) != {None} and set(plan_dict.values()) != {0}:
                update_csv(output_dir, headers, plan_values_dict)

            fact_values_dict = {
                'Ресурсы': resource_name + '_res',
                'Субподрядчик': cell.value,
                'план/факт': 'факт',
            }

            fact_dict = dict(zip(days, fact_list))
            fact_values_dict.update(fact_dict)

            # только пустые значения в строке не добавляются в csv
            print(fact_dict)
            if set(fact_dict.values()) != {None} and set(fact_dict.values()) != {0}:
                print('зашел')
                update_csv(output_dir, headers, fact_values_dict)

def main():
    inp_file_path = 'data/5.xlsm'
    wb = load_workbook(inp_file_path, data_only=True)

    activities_sheet = wb['МСГ']
    resources_sheet = wb['Ресурсы']

    get_activites_data(activities_sheet, 30)
    get_resources_data(resources_sheet, 30)

if __name__ == "__main__":
    main()
